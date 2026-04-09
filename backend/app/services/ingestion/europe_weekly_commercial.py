"""EUROPE Weekly Commercial Dashboard Sheet1 layout — Phase 7 (value rows only, D2)."""

from __future__ import annotations

import io
import re
from zipfile import BadZipFile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

from app.services.ingestion.validator import ValidationResult, ValidatedRow

# Sheet1 row 2 = headers; col B = Sr. No., C/D = customer legal/common; E+ = month buckets.
_MIN_COLS = 5


def looks_like_europe_weekly_sheet(rows: list[tuple[Any, ...]]) -> bool:
    """True if row 2 (index 1) resembles the reference workbook header row."""
    if len(rows) < 2:
        return False
    r2 = rows[1]
    if not r2 or len(r2) < _MIN_COLS:
        return False
    cells = [str(c).strip().lower() if c is not None else "" for c in r2[:10]]
    joined = " ".join(cells)
    if "customer" not in joined:
        return False
    compact = re.sub(r"[\s._]", "", joined)
    return "sr" in compact and "no" in compact


def try_parse_europe_weekly_commercial(content: bytes) -> ValidationResult | None:
    """If Sheet1 matches the EUROPE layout, return validation result; else None (use Phase 1 path)."""
    bio = io.BytesIO(content)
    try:
        wb = load_workbook(filename=bio, read_only=True, data_only=True)
    except BadZipFile:
        return None
    try:
        if "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
        else:
            ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not looks_like_europe_weekly_sheet(rows):
        return None

    header = rows[1]
    month_starts: list[date] = []
    for c in header[4:]:
        ms = _cell_to_month_start(c)
        if ms is not None:
            month_starts.append(ms)
    if not month_starts:
        return ValidationResult(
            errors=[
                {
                    "row": 2,
                    "column": None,
                    "message": "No month columns found after customer names (expected dates from column E).",
                }
            ],
            rows=[],
        )

    errors: list[dict[str, Any]] = []
    pending: list[ValidatedRow] = []

    for ir in range(2, len(rows)):
        raw = rows[ir]
        excel_row = ir + 1
        if raw is None:
            continue
        row = list(raw)
        while len(row) < len(header):
            row.append(None)

        sr = row[1] if len(row) > 1 else None
        sr_empty = sr is None or (isinstance(sr, str) and not str(sr).strip())
        if sr_empty:
            continue

        legal = _opt_str(row[2] if len(row) > 2 else None)
        common = _opt_str(row[3] if len(row) > 3 else None)
        if not legal:
            continue

        for j, month_start in enumerate(month_starts):
            col_idx = 4 + j
            if col_idx >= len(row):
                break
            cell = row[col_idx]
            if cell is None or cell == "":
                continue
            try:
                if isinstance(cell, bool):
                    continue
                amt = Decimal(str(cell).strip().replace(",", ""))
            except (InvalidOperation, ValueError):
                errors.append(
                    {
                        "row": excel_row,
                        "column": f"col_{col_idx + 1}",
                        "message": "Amount must be a decimal number.",
                    }
                )
                continue
            if amt <= 0:
                continue

            row_index = excel_row * 1000 + col_idx
            pending.append(
                ValidatedRow(
                    row_index=row_index,
                    amount=amt,
                    revenue_date=month_start,
                    business_unit=None,
                    division=None,
                    customer=legal,
                    revenue_type=None,
                    customer_name_common=common,
                )
            )

    if errors:
        return ValidationResult(errors=errors, rows=[])
    validated = pending
    if not validated:
        return ValidationResult(
            errors=[
                {
                    "row": None,
                    "column": None,
                    "message": "No value rows with positive amounts found for EUROPE Weekly Commercial layout.",
                }
            ],
            rows=[],
        )
    return ValidationResult(errors=[], rows=validated)


def _opt_str(v: object | None) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _cell_to_month_start(value: object | None) -> date | None:
    """Normalize header cell to first day of month (revenue month bucket)."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    s = str(value).strip()
    try:
        d = datetime.fromisoformat(s.replace("Z", "+00:00")[:10])
        return date(d.year, d.month, 1)
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            d = datetime.strptime(s[:10], fmt)
            return date(d.year, d.month, 1)
        except ValueError:
            continue
    return None
