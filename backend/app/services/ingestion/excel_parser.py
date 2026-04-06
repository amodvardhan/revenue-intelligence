"""Parse Excel bytes into tabular rows for validation and preview."""

from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import load_workbook


@dataclass(frozen=True)
class ParsedExcel:
    """First sheet: header row + data rows."""

    headers: list[str]
    rows: list[list[object | None]]
    sheet_name: str


def parse_excel(content: bytes) -> ParsedExcel:
    """Load the first worksheet; row 1 is headers, row 2+ are data."""
    bio = io.BytesIO(content)
    wb = load_workbook(filename=bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        sheet_name = ws.title
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return ParsedExcel(headers=[], rows=[], sheet_name=sheet_name)
        headers = [_normalize_header(h) for h in header_row]
        data_rows: list[list[object | None]] = []
        for row in rows_iter:
            if row is None:
                continue
            data_rows.append(list(row))
        return ParsedExcel(headers=headers, rows=data_rows, sheet_name=sheet_name)
    finally:
        wb.close()


def _normalize_header(value: object | None) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower().replace(" ", "_")
    return s
